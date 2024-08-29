from django.shortcuts import render, redirect,  get_object_or_404
from .forms import *
from django.contrib import messages
from datetime import datetime, date
from .functions import *
from proyectos.models import Proyectos
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
# cambio esto para probar un modelo de usuario con una opción OneToOneField con User
from .models import registro, Reporte_tarea, AusenciaJustificada  # User,
from django.contrib.auth.models import User, Group
from django.http import JsonResponse
import openpyxl
from io import BytesIO
from django.http import HttpResponse
import calendar
from django.utils.dateparse import parse_date
from django.core.files.storage import default_storage
from django.conf import settings

app_name = 'asistencias'
# Create your views here.

# Vista que renderiza el menú asistencia. Verifica si el usuario tiene un permiso para agregar registros
# y en función de eso modifica un booleano para mostrar o no ciertos elementos del template.
# Arma el contexto necesario en el diccionario data y lo pasa al render.


@login_required
def asistencia(request):
    crear = False
    if request.user.has_perm('asistencias.add_registro'):
        crear = True
    usuario = User.objects.get(username=request.user).get_full_name()
    img = User.objects.get(username=request.user).usuario.image.url
    data = {
        'usuario': usuario,
        'img': img,
        'crear': crear,
    }
    return render(request, 'asistencias/asistencia.html', data)

# Vista que le permite al usuario calcular el tiempo que lleva trabajado durante la jornada.
# Cabe aclarar que por la lógica de la aplicación, el cálculo se hace entre una marca de entrada y la hora actual.


@login_required
def hoy(request):
    # usuario = User.objects.get(username=request.user).get_full_name()
    # id = User.objects.get(username=request.user).id
    # fecha = date.today().strftime("%Y-%m-%d")
    # registros = registro.objects.filter(fecha=fecha).filter(usuario=id)
    # hora_actual = datetime.today().time()
    # hora_actual_obj = datetime.strptime(
    #     str(hora_actual.hour)+':'+str(hora_actual.minute), '%H:%M')
    # if len(registros) > 0:
    #     tiempo = calcular_horas_totales_hoy(registros, hora_actual_obj)
    #     mensaje = f'Usted trabajó {tiempo} horas el dia de hoy.'
    # else:
    #     mensaje = 'Todavía no se ha registrado una entrada para el día de hoy.'
    # data = {
    #     'usuario': usuario,
    #     'fecha': fecha,
    #     'mensaje': mensaje,
    # }
    usuario = User.objects.get(username=request.user).get_full_name()
    data = {'usuario': usuario}
    return render(request, 'core/disculpe_las_molestias.html', data)

# Vista que permite calcular las horas trabajadas en un rango de fechas. La vista está pensada para renderizar
# el template inicial y para procesar la devolución. Prepara el contexto en la variable data en donde lleva un booleano
# que indica si el usuario es capaz de calcular las horas de cualquier persona o si solo puede calcular las suyas.


@login_required
def ver_periodo(request):
    # Usus = User.objects.all()
    # username = request.user
    # if request.method == 'POST':
    #     usuario = request.POST['usuario']
    #     u_id = User.objects.get(username=usuario)
    #     fecha1_str = request.POST['fecha1']
    #     fecha1_obj = datetime.strptime(fecha1_str, '%Y-%m-%d')
    #     fecha2_str = request.POST['fecha2']
    #     fecha2_obj = datetime.strptime(fecha2_str, '%Y-%m-%d')
    #     mensaje = []
    #     mensaje = Calcular_horas(fecha1_obj, fecha2_obj, u_id)
    #     data = {
    #         'usuario': usuario,
    #         'username': username,
    #         'Usus': Usus,
    #     }
    #     if mensaje[1] == True:
    #         messages.error(request, mensaje[0])
    #     else:
    #         messages.success(request, mensaje[0])
    #     return render(request, 'asistencias/ver_periodo.html', data)

    # usuario = User.objects.get(username=username).get_full_name()
    # autorizado = False
    # grupos_con_permiso = ['Gerencia', 'Jefe Area']
    # if any(username.groups.filter(name=grupo).exists() for grupo in grupos_con_permiso):
    #     autorizado = True
    # data = {
    #     'usuario': usuario,
    #     'username': username,
    #     'Usus': Usus,
    #     'autorizado': autorizado,
    # }
    # return render(request, 'asistencias/ver_periodo.html', data)
    usuario = User.objects.get(username=request.user).get_full_name()
    data = {'usuario': usuario}
    return render(request, 'core/disculpe_las_molestias.html', data)

# Registros

# Vista que permite al usuario registrar SU entrada o salida. No puede cargar la de otro usuario.


@login_required
def marcar_asistencia(request):
    desdeRegistros = False
    desdeHoy = False
    if 'verAsistencias' in request.META.get('HTTP_REFERER'):
        desdeRegistros = True
    if 'hoy' in request.META.get('HTTP_REFERER'):
        desdeHoy = True
    if request.method == 'POST':
        try:
            usuario = request.POST['usuario']
            fecha_str = request.POST['fecha']
            fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d')
            hora = request.POST['hora']
            tipo = request.POST['tipo']
            registro.objects.create(
                usuario=User.objects.get(username=usuario),
                fecha=fecha_obj,
                hora=hora,
                tipo=tipo
            )
            messages.success(request, 'Asistencia registrada con exito.')
        except Exception as e:
            messages.error(
                request, 'Hubo un error al registrar la asistencia: ' + str(e) + '.')
        return redirect('verAsistencias')
    Usus = User.objects.all()
    username = request.user
    autorizado = False
    if str(username) == 'superusuario':
        autorizado = True
    usuario = User.objects.get(username=username).get_full_name()
    user_id = User.objects.get(username=username).id
    fecha_guion = date.today().strftime("%Y-%m-%d")
    hora = datetime.now().time()
    # obtengo el último registro del usuario en el día de la fecha de modo de que el formulario induzca a marcar una entrada o salida según que sea lo último que se registró.
    try:
        ult_reg = registro.objects.filter(
            fecha=fecha_guion).filter(usuario=user_id).last().tipo
    except:
        ult_reg = None
    data = {
        'usuario': usuario,
        'username': username,
        'Usus': Usus,
        'fecha': fecha_guion,
        'hora': hora,
        'ult_reg': ult_reg,
        'autorizado': autorizado,
        'desdeRegistros': desdeRegistros,
        'desdeHoy': desdeHoy,
    }
    return render(request, 'asistencias/marcar_asistencia.html', data)


@csrf_exempt
def enviar_asistencia(request):
    if request.method == 'POST':
        try:
            # Separar y procesar los datos recibidos

            datos = request.POST['datos_a_enviar'].split('|')
            nombre = datos[0].split(' ')[0]
            print("nombre: ", nombre)
            user = User.objects.get(first_name=nombre)
            tipo = datos[1]
            fecha_str = datos[2].split(' ', 1)[1]
            fecha = datetime.strptime(
                fecha_str, '%d/%m/%Y')
            hora_str = datos[3]
            hora = datetime.strptime(hora_str, '%H:%M').time()

            # Crear el registro en la base de datos
            registro.objects.create(
                usuario=user,
                fecha=fecha,
                hora=hora,
                tipo=tipo
            )

            # Responder con un JSON y un código 200
            return JsonResponse({'status': 'success', 'message': 'Asistencia registrada con éxito.'}, status=200)
        except Exception as e:
            # Manejar errores y responder con un código 500
            print(f"Error en enviar_asistencia: {e}")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    else:
        return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)

# Vista para ver todos los registros de asistencias de todos los usuarios.
# Se evalua los permisos del usuario que hace la solicitud para habilitar ciertas funciones
# o incluso ocultar algunos elementos del template.


@login_required
def listar_registros(request):
    crear = False
    editar = False
    borrar = False
    autorizado = False
    if request.user.groups.filter(name='Jefe Area').exists() or str(request.user) == 'superusuario':
        autorizado = True
    if request.user.has_perm('asistencias.add_registro'):
        crear = True
    if request.user.has_perm('asistencias.change_registro'):
        editar = True
    if request.user.has_perm('asistencias.delete_registro'):
        borrar = True

    usuario = User.objects.get(username=request.user).get_full_name()
    registros = registro.objects.all()
    data = {
        'usuario': usuario,
        'registros': registros,
        'crear': crear,
        'borrar': borrar,
        'editar': editar,
        'autorizado': autorizado,
    }
    return render(request, 'asistencias/registros.html', data)

# Vista que permite la edición de un registro de asistencia. Esta vista siempre es
# llamada por un formulario que trae un POST con la información por eso no lleva la verificación.


@login_required
def editar_registro(request):
    try:
        reg = registro.objects.get(id=request.POST['id'])
        cambios = False
        if request.POST['fecha'] != reg.fecha:
            reg.fecha = request.POST['fecha']
            cambios = True
        if request.POST['hora'] != reg.hora:
            reg.hora = request.POST['hora']
            cambios = True
        if request.POST['tipo'] != reg.tipo:
            reg.tipo = request.POST['tipo']
            cambios = True
        if cambios:
            reg.save()
            messages.success(request, "Registro editado con exito.")
    except Exception as e:
        messages.error(
            request, "Hubo un error al editar el registro: " + str(e) + ".")
    return redirect('verAsistencias')

# Vista que permite el borrado de un registro de asistencia.


@login_required
def borrar_registro(request, id):
    try:
        reg = registro.objects.get(id=id)
        reg.delete()
        messages.success(request, "Registro borrado con exito.")
    except Exception as e:
        messages.error(
            request, "Hubo un error al borrar el registro: " + str(e) + ".")
    return redirect('verAsistencias')

# Menu Reportes
# Se evalua el permiso de creación para ocultar el botón de crear un nuevo reporte.


@login_required
def reportes(request):
    crear = False
    if request.user.has_perm('asistencias.add_registro'):
        crear = True
    usuario = User.objects.get(username=request.user).get_full_name()
    img = User.objects.get(username=request.user).usuario.image.url
    data = {
        'usuario': usuario,
        'img': img,
        'crear': crear,
    }
    return render(request, 'asistencias/reportes.html', data)

# Vista que lista todos los reportes de las tareas realizadas en el día por el personal.
# Se evaluan los permisos del usuario que ingresa a la vista de modo de limitar ciertas acciones como editar o borrar registros
# o incluso la creación de uno nuevo.


@login_required
def ver_reportes(request):
    crear = False
    editar = False
    borrar = False
    autorizado = False
    if request.user.groups.filter(name='Jefe Area').exists() or str(request.user) == 'superusuario':
        autorizado = True
    if request.user.has_perm('asistencias.add_reporte_tarea'):
        crear = True
    if request.user.has_perm('asistencias.change_reporte_tarea'):
        editar = True
    if request.user.has_perm('asistencias.delete_reporte_tarea'):
        borrar = True
    usuario = User.objects.get(username=request.user).get_full_name()
    reportes = Reporte_tarea.objects.all()
    data = {
        'usuario': usuario,
        'reportes': reportes,
        'crear': crear,
        'borrar': borrar,
        'editar': editar,
        'autorizado': autorizado,
    }
    return render(request, 'asistencias/ver_reportes.html', data)

# Vista para crear reportes. Está pensada para procesar la creación de un nuevo reporte y
# recibir la información una vez que se completa el formulario.


@login_required
def crear_reporte(request):
    desdeReportes = False
    if 'verReportes' in request.META.get('HTTP_REFERER'):
        desdeReportes = True
    if request.method == 'POST':
        # if request.POST["usuario"] == request.user and request.POST['fecha'] <= date.today().strftime("%Y-%m-%d") and request.POST['hora'] <= datetime.now().time():
        try:
            Reporte_tarea.objects.create(
                usuario=User.objects.get(username=request.POST['usuario']),
                informe=request.POST['informe'],
                fecha=request.POST['fecha'],
                hora=request.POST['hora'],
            )
            if request.POST['proyecto'] != 'Ninguno':
                r = Reporte_tarea.objects.last()
                r.proyecto = Proyectos.objects.get(
                    nombre=request.POST['proyecto'])
                r.save()
            messages.success(request, 'Reporte creado con exito.')
        except Exception as e:
            messages.error(
                request, 'Hubo un error al crear el reporte: ' + str(e) + '.')

        return redirect('verReportes')
    else:
        nombre_usuario = request.user
        usuario = User.objects.get(username=request.user).get_full_name()
        fecha = date.today().strftime("%Y-%m-%d")
        hora = datetime.now().time()
        Proys = Proyectos.objects.all()
        data = {
            'usuario': usuario,
            'nombre_usuario': nombre_usuario,
            'fecha': fecha,
            'hora': hora,
            'Proys': Proys,
            'desdeReportes': desdeReportes,
        }
        return render(request, 'asistencias/nuevo_reporte.html', data)

# Vista para editar un reporte diario. Esta vista siempre es
# llamada por un formulario que trae un POST con la información por eso no lleva la verificación.


@login_required
def editar_reporte(request):
    try:
        r = Reporte_tarea.objects.get(id=request.POST['id'])
        cambios = False
        if request.POST['informe'] != r.informe:
            r.informe = request.POST['informe']
            cambios = True
        if request.POST['fecha'] != r.fecha:
            r.fecha = request.POST['fecha']
            cambios = True
        if request.POST['hora'] != r.hora:
            r.hora = request.POST['hora']
            cambios = True
        if request.POST['proyecto'] == 'Ninguno':
            r.proyecto = None
        else:
            if request.POST['proyecto'] != r.proyecto:
                r.proyecto = Proyectos.objects.get(
                    nombre=request.POST['proyecto'])
                cambios = True
        if cambios:
            r.save()
            messages.success(request, 'Reporte editado con exito.')
    except Exception as e:
        messages.error(
            request, 'Hubo un error al editar el reporte: ' + str(e) + '.')
    return redirect('verReportes')

# Vista que permite el borrado de un reporte.


@login_required
def borrar_reporte(request, id):
    try:
        r = Reporte_tarea.objects.get(id=id)
        r.delete()
        messages.success(request, 'Reporte borrado con exito.')
    except Exception as e:
        messages.error(
            request, 'Hubo un error al borrar el reporte: ' + str(e) + '.')
    return redirect('verReportes')


# Vista que recibe una petición desde el frontend y permite generar un excel con los registros de asistencia entre dos fechas seleccinoadas.
@login_required
def exportar_registros(request):
    if request.method == 'POST':
        todos = request.POST.get('todos')
        fecha1 = request.POST.get('fecha1')
        fecha2 = request.POST.get('fecha2')

        # Convertir las fechas desde la cadena al formato datetime
        if fecha1:
            fecha1 = parse_date(fecha1)
        if fecha2:
            fecha2 = parse_date(fecha2)

        # Crear un libro de trabajo
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Eliminar la hoja inicial

        # Filtrar registros según las fechas si se proporcionan
        registros = registro.objects.all()
        if todos == 'true':  # Verificar si se solicitó todos los registros
            print("Descargando todos los registros.")
        else:
            print("Descargando registros entre fechas.")
            if fecha1 and fecha2:
                registros = registros.filter(fecha__range=[fecha1, fecha2])
            elif fecha1:
                registros = registros.filter(fecha__gte=fecha1)
            elif fecha2:
                registros = registros.filter(fecha__lte=fecha2)

        # Depuración: número de registros filtrados
        print(f"Número de registros encontrados: {registros.count()}")

        # Agrupar registros por año y mes
        registros_por_mes = {}
        for reg in registros:
            key = (reg.fecha.year, reg.fecha.month)
            if key not in registros_por_mes:
                registros_por_mes[key] = []
            registros_por_mes[key].append(reg)

        # Crear hojas por mes con registros
        for (year, month), registros in registros_por_mes.items():
            nombre_mes = f"{calendar.month_name[month]} {year}"
            sheet = wb.create_sheet(title=nombre_mes)

            # Definir encabezados
            headers = ['Número', 'Nombre', 'Tipo', 'Fecha', 'Hora']
            sheet.append(headers)

            # Agregar datos de los registros
            for reg in registros:
                data = [
                    reg.id,
                    reg.usuario.get_full_name() if reg.usuario else 'Desconocido',
                    reg.tipo,
                    reg.fecha.strftime('%Y-%m-%d'),
                    reg.hora.strftime('%H:%M:%S')
                ]
                sheet.append(data)

        # Guardar el libro en un buffer de memoria
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Crear la respuesta HTTP
        response = HttpResponse(
            buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=registros_asistencia.xlsx'

        return response


# Vista que recibe una petición desde el frontend y permite generar un excel con los registros de asistencia entre dos fechas seleccinoadas.
@login_required
def exportar_reportes(request):
    if request.method == 'POST':
        todos = request.POST.get('todos')
        fecha1 = request.POST.get('fecha1')
        fecha2 = request.POST.get('fecha2')

        # Convertir las fechas desde la cadena al formato datetime
        if fecha1:
            fecha1 = parse_date(fecha1)
        if fecha2:
            fecha2 = parse_date(fecha2)

        # Crear un libro de trabajo
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Eliminar la hoja inicial

        # Filtrar reportes según las fechas si se proporcionan
        reportes = Reporte_tarea.objects.all()
        if todos == 'true':  # Verificar si se solicitó todos los reportes
            print("Descargando todos los reportes.")
        else:
            print("Descargando reportes entre fechas.")
            if fecha1 and fecha2:
                reportes = reportes.filter(fecha__range=[fecha1, fecha2])
            elif fecha1:
                reportes = reportes.filter(fecha__gte=fecha1)
            elif fecha2:
                reportes = reportes.filter(fecha__lte=fecha2)

        # Depuración: número de reportes filtrados
        print(f"Número de reportes encontrados: {reportes.count()}")

        # Agrupar reportes por año y mes
        reportes_por_mes = {}
        for reg in reportes:
            key = (reg.fecha.year, reg.fecha.month)
            if key not in reportes_por_mes:
                reportes_por_mes[key] = []
            reportes_por_mes[key].append(reg)

        # Crear hojas por mes con reportes
        for (year, month), reportes in reportes_por_mes.items():
            nombre_mes = f"{calendar.month_name[month]} {year}"
            sheet = wb.create_sheet(title=nombre_mes)

            # Definir encabezados
            headers = ['Número', 'Nombre', 'Informe',
                       'Proyecto', 'Fecha', 'Hora']
            sheet.append(headers)

            # Agregar datos de los reportes
            for rep in reportes:
                data = [
                    rep.id,
                    rep.usuario.get_full_name() if rep.usuario else 'Desconocido',
                    rep.informe,
                    rep.proyecto.nombre if rep.proyecto else 'Ninguno',
                    rep.fecha,
                    rep.hora
                ]
                sheet.append(data)

        # Guardar el libro en un buffer de memoria
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Crear la respuesta HTTP
        response = HttpResponse(
            buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=reportes_asistencia.xlsx'

        return response

# Vista que permite visualizar una tabla con las ausencias registradas.


@login_required
def ver_ausencias(request):
    usuario = User.objects.get(username=request.user).get_full_name()
    autorizado = False
    if request.user.groups.filter(name__in=['Jefe Area', 'Equipo Administrativo']).exists() or str(request.user) == 'superusuario':
        autorizado = True
    Aus = AusenciaJustificada.objects.all()
    data = {
        'usuario': usuario,
        'Aus': Aus,
        'autorizado': autorizado,
    }
    return render(request, 'asistencias/ausencias.html', data)

# Vista llamada desde el frontend para obtener información relevante a la hora de crear o editar una ausencia.


@login_required
def obtener_ausencia(request, ausencia_id):
    autorizado = request.user.groups.filter(
        name__in=['Jefe Area', 'Equipo Administrativo']).exists() or request.user.is_superuser

    if ausencia_id == 'new':
        data = {
            'autorizado': autorizado,
            'usuario': request.user.username
        }
        if autorizado:
            usuarios = User.objects.all().values('id', 'username')
            data['usuarios'] = list(usuarios)
    else:
        try:
            ausencia = AusenciaJustificada.objects.get(id=int(ausencia_id))
        except AusenciaJustificada.DoesNotExist:
            return JsonResponse({'error': 'El registro de ausencia no existe.'}, status=404)

        puede_editar = autorizado or ausencia.usuario == request.user
        data = {
            'fecha_inicio': ausencia.fecha_inicio,
            'fecha_fin': ausencia.fecha_fin,
            'cantidad_dias': ausencia.cantidad_dias,
            'motivo': ausencia.motivo,
            'archivo': ausencia.archivo.url if ausencia.archivo else '',
            'autorizado': autorizado,
            'usuario': ausencia.usuario.username,
            'puede_editar': puede_editar,
            'archivo_existe': bool(ausencia.archivo),
        }
        if autorizado:
            usuarios = User.objects.all().values('id', 'username')
            data['usuarios'] = list(usuarios)

    return JsonResponse(data)


@login_required
def guardar_ausencia(request):
    if request.method == "POST":
        ausencia_id = request.POST.get('id')
        usuario_select = request.POST.get('usuarioSelect')
        usuario_input = request.POST.get('usuarioInput')
        usuario = usuario_select if usuario_select else usuario_input
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        motivo = request.POST.get('motivo')
        cantidad_dias = request.POST.get('cantidad_dias')
        archivo = request.FILES.get('archivo', None)
        archivos_eliminados = request.POST.getlist('archivos_eliminados[]')

        # Mensajes de depuración para verificar los datos recibidos
        print(f"ID Ausencia: {ausencia_id}")
        print(f"Usuario: {usuario}")
        print(f"Fecha Inicio: {fecha_inicio}")
        print(f"Fecha Fin: {fecha_fin}")
        print(f"Motivo: {motivo}")
        print(f"Cantidad Días: {cantidad_dias}")
        print(f"Archivo: {archivo}")
        print(f"Archivos Eliminados: {archivos_eliminados}")

        if ausencia_id == 'new':
            ausencia = AusenciaJustificada(
                usuario_id=User.objects.get(id=usuario).id,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                motivo=motivo,
                cantidad_dias=cantidad_dias if cantidad_dias else None,
                archivo=archivo
            )
            ausencia.save()
            messages.success(
                request, "Registro de ausencia creado exitosamente.")

        else:
            ausencia = get_object_or_404(
                AusenciaJustificada, id=ausencia_id)

            if (ausencia.usuario == request.user or
                request.user.groups.filter(name__in=['Jefe Area', 'Equipo administrativo']).exists() or
                    request.user.is_superuser):

                ausencia.fecha_inicio = fecha_inicio
                ausencia.fecha_fin = fecha_fin
                ausencia.motivo = motivo
                ausencia.cantidad_dias = cantidad_dias if cantidad_dias else None

                # Eliminar archivos del servidor
                for archivo_url in archivos_eliminados:
                    print(f"Archivo URL: {archivo_url}")

                    # Obtener el path relativo al directorio MEDIA_ROOT
                    archivo_path = archivo_url.lstrip('/media/')
                    print(f"Archivo Path: {archivo_path}")

                    full_path = os.path.join(
                        settings.MEDIA_ROOT, archivo_path)
                    print(f"Full Media Path: {full_path}")
                    print(
                        f"File Exists: {default_storage.exists(archivo_path)}")

                    print(
                        f"Intentando eliminar el archivo: {archivo_path}")
                    if default_storage.exists(archivo_path):
                        default_storage.delete(archivo_path)
                        ausencia.archivo = None
                        print(f"Archivo eliminado: {archivo_path}")
                    else:
                        print(f"Archivo no encontrado: {archivo_path}")

                # Actualizar archivo si se sube uno nuevo
                if archivo:
                    ausencia.archivo = archivo
                    print(f"Archivo actualizado: {archivo}")

                ausencia.save()
                messages.success(
                    request, "Registro de ausencia editado exitosamente.")
            else:
                messages.error(
                    request, "No tienes permisos para editar este registro de ausencia.")

        return redirect('/asistencias/verAusencias')

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def borrar_ausencia(request, id):
    try:
        ausencia = AusenciaJustificada.objects.get(id=id)
        ausencia.delete()
        messages.success(request, 'Registro de ausencia borrado con exito.')
    except Exception as e:
        messages.error(
            request, 'Hubo un error al borrar el registro: ' + str(e) + '.')
    return redirect('verAusencias')


@login_required
def eliminar_archivo(request, ausencia_id):
    if request.method == 'POST':
        try:
            ausencia = AusenciaJustificada.objects.get(id=int(ausencia_id))
            if ausencia.archivo:
                # Elimina el archivo del sistema de archivos
                ausencia.archivo.delete(save=False)
                ausencia.archivo = None
                ausencia.save()
            return JsonResponse({'success': True})
        except AusenciaJustificada.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'El registro de ausencia no existe.'}, status=404)
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
