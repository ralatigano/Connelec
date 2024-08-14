from django.shortcuts import render, redirect
from .models import Tareas, Archivos, Entrada_historial
from django.contrib import messages
from proyectos.models import Proyectos
from datetime import datetime
from django.http import JsonResponse
from django.core import serializers
from .functions import *
from urllib.parse import urlparse
import openpyxl
from io import BytesIO
from django.contrib.auth.decorators import login_required
# cambio esto para probar un modelo de usuario con una aplicación OneToOneField con User
from django.contrib.auth.models import User
# from asistencias.models import User

app_name = 'tareas'

# Create your views here.

# Menu tareas.


@login_required
def tareas(request):
    global desdeProyectos
    desdeProyectos = False
    usuario = User.objects.get(username=request.user).get_full_name()
    img = User.objects.get(username=request.user).usuario.image.url
    data = {
        'usuario': usuario,
        'img': img,
    }
    return render(request, 'tareas/tareas.html', data)

# Vista que enlista todas las tareas registradas. Por una cuestión de funcionalidad, la posibilidad de editar/borrar estos registros se
# va a manejar desde el frontend con javascript según que usuarios haya creado las tareas o quien esté designado como encargado.


@login_required
def ver_tareas(request):
    usuario = User.objects.get(username=request.user).get_full_name()
    Tars = Tareas.objects.all().order_by('-fecha_creacion')
    data = {
        'usuario': usuario,
        'Tars': Tars,
    }
    return render(request, 'tareas/ver_tareas.html', data)

# Vista que permite la creación de una nueva tarea. La vista está pensada para mostrar el formulario de creación y para manejar el POST que resulta de
# de completarlo y guardarlo.


@login_required
def crear_tarea(request):
    # banderas para saber desde que vista estoy llegando a crear la tarea
    desdeVerTareas = False
    desdeMisTareas = False
    nombreProyecto = ''
    desdeTareas = False
    global desdeProyectos
    if desdeProyectos:
        nombreProyecto = request.META.get('HTTP_REFERER').split('/')[5]
        nombreProyecto = nombreProyecto.replace('%20', ' ')
    if 'verTareas' in request.META.get('HTTP_REFERER'):
        desdeVerTareas = True
    if 'MisTareas' in request.META.get('HTTP_REFERER'):
        desdeMisTareas = True
    if not desdeVerTareas and desdeMisTareas and desdeProyectos:
        desdeTareas = True
    usuario = User.objects.get(username=request.user).get_full_name()
    Usus = User.objects.all()
    Proys = Proyectos.objects.all()
    # creo la tarea
    if request.method == 'POST':
        try:
            Tareas.objects.create(
                nombre=request.POST['nombre'],
                descrip=request.POST['descripcion'],
                fecha_entrega=None,
                estado=request.POST['estado'],
                proyecto=None,
                encargado=None,
                created_by=request.user,
            )
            if request.POST['fecha_entrega'] == 'dd/mm/aaaa':
                tarea = Tareas.objects.last()
                tarea.fecha_entrega = request.POST['fecha_entrega']
                tarea.save()
            if request.POST['proyecto'] != 'Ninguno':
                tarea = Tareas.objects.last()
                tarea.proyecto = Proyectos.objects.get(
                    nombre=request.POST['proyecto'])
                tarea.save()
            if request.POST['encargado'] != 'Ninguno':
                tarea = Tareas.objects.last()
                tarea.encargado = User.objects.get(
                    username=request.POST['encargado'])
                tarea.save()
                # Si hay un asignado, me interesa notificarlo.
                usuario = request.user
                tarea_id = tarea.id

                notificar_encargado(tarea_id, usuario)
            if request.FILES:
                for archivo in request.FILES.getlist('adjunto'):
                    Archivos.objects.create(
                        nombre=archivo.name,
                        tarea=Tareas.objects.last(),
                        archivo=archivo
                    )

            messages.success(request, 'Tarea creada con exito.')
        except Exception as e:
            messages.error(
                request, 'Hubo un error al crear la tarea: ' + str(e) + '.')
        # ve adonde redirijo
        if desdeProyectos:
            desdeProyectos = False
            tarea = Tareas.objects.last()
            return redirect('porProyecto/' + tarea.proyecto.nombre)
        elif desdeVerTareas:
            return redirect('verTareas')
        elif desdeMisTareas:
            tarea = Tareas.objects.last()
            return redirect('verMisTareas')
        else:
            return redirect('crearTarea')
    data = {
        'usuario': usuario,
        'Usus': Usus,
        'Proys': Proys,
        'desdeProyectos': desdeProyectos,
        'nombreProyecto': nombreProyecto,
        'desdeVerTareas': desdeVerTareas,
        'desdeMisTareas': desdeMisTareas,
    }
    return render(request, 'tareas/crear_tarea.html', data)

# Vista que se usa a traves de una petición AJAX desde el frontend para obtener la información de los usuarios y de los proyectos en formato JSON.
# y utilizarla para completar un modal de edición de tareas.


@login_required
def info_editar_tarea(request):
    Usus = User.objects.all()
    Proys = Proyectos.objects.all()

    Usus_data = list(Usus.values('id', 'username', 'first_name', 'last_name'))
    Proys_data = list(Proys.values('nombre'))
    data = {
        'Usus': Usus_data,
        'Proys': Proys_data
    }
    return JsonResponse(data)

# Vista que recibe el POST del frontend con la información para editar una tarea.


@login_required
def editar_tarea(request):

    try:
        tarea = Tareas.objects.get(nombre=request.POST['nombre'])
        cambios = False
        if tarea.descrip != request.POST['descrip'] and request.POST['descrip'] != '':
            tarea.descrip = request.POST['descrip']
            cambios = True
        if tarea.encargado != request.POST['encargado']:
            if request.POST['encargado'] == 'Ninguno':
                tarea.encargado = None
            else:
                tarea.encargado = User.objects.get(
                    username=request.POST['encargado'])
            cambios = True
        if tarea.proyecto != request.POST['proyecto']:
            if request.POST['proyecto'] == 'Ninguno':
                tarea.proyecto = None
            else:
                tarea.proyecto = Proyectos.objects.get(
                    nombre=request.POST['proyecto'])
            cambios = True
        if tarea.estado != request.POST['estado']:
            tarea.estado = request.POST['estado']
            cambios = True
        if tarea.fecha_entrega != request.POST['fecha_entrega']:
            if request.POST['fecha_entrega'] == '':
                tarea.fecha_entrega = None
            else:
                tarea.fecha_entrega = request.POST['fecha_entrega']
            cambios = True
        if request.FILES:
            for archivo in request.FILES.getlist('adjunto'):
                Archivos.objects.create(
                    nombre=archivo.name,
                    tarea=Tareas.objects.get(id=tarea.id),
                    archivo=archivo
                )
            cambios = True
        if cambios:
            tarea.save()
            messages.success(request, 'Tarea editada con exito.')
    except Exception as e:
        messages.error(
            request, 'Hubo un error al editar la tarea: ' + str(e) + '.')
    return redirect('verTareas')

# Vista que permite borrar un registro de una tarea según su ID.


@login_required
def borrar_tarea(request, id):
    try:
        Tareas.objects.get(id=id).delete()
        messages.success(request, "Tarea borrada con exito.")
    except Exception as e:
        messages.error(
            request, "Hubo un error al borrar la tarea: " + str(e) + ".")
    return redirect('verTareas')


# Vistas particulares de las tareas, filtradas por proyecto o por encargado o vista de una tarea en particular.
# Vista de tareas filradas por el request.user
@login_required
def ver_mis_tareas(request):

    usuario = User.objects.get(username=request.user).get_full_name()
    Tars = Tareas.objects.filter(
        encargado=request.user).order_by('-fecha_creacion')
    data = {
        'usuario': usuario,
        'Tars': Tars,
    }
    return render(request, 'tareas/ver_mis_tareas.html', data)

# Vista que filtra las tareas vinculadas a un proyecto en particular.


@login_required
def tareas_asosc_proy(request, proy):
    global desdeProyectos
    desdeProyectos = True
    usuario = User.objects.get(username=request.user).get_full_name()
    Usus = User.objects.all()
    p = Proyectos.objects.get(nombre=proy).id
    Tars = Tareas.objects.filter(proyecto=p)
    data = {
        'usuario': usuario,
        'Usus': Usus,
        'Tars': Tars,
        'proyecto': proy,
    }
    return render(request, 'tareas/tareas_asoc_proyecto.html', data)

# Vista que permite ver los detalles de una tarea en particular.


@login_required
def explorar_tarea(request, nombTarea):
    usuario = User.objects.get(username=request.user).get_full_name()
    Usus = User.objects.all()
    try:
        Tarea = Tareas.objects.get(nombre=nombTarea)
        Files = Archivos.objects.filter(tarea=Tarea.id)
        data = {
            'usuario': usuario,
            'Usus': Usus,
            'Tarea': Tarea,
            'Files': Files
        }
    except Exception as e:
        messages.error(
            request, 'Hubo un error al recuperar la información de la tarea: ' + str(e) + '.')
    return render(request, 'tareas/explorar.html', data)

# Vista que permite ver las entradas historicas relacionadas a un proyecto particular. Se evalúan los permisos
# de usuario para habilitar o no ciertas funciones o incluso mostrar/ocultar algunos elementos del template.


@login_required
def entradas_asosc_proy(request, proy):
    autorizado = False
    if request.user.groups.filter(name='Jefe Area').exists() or str(request.user) == 'superusuario':
        autorizado = True
    usuario = User.objects.get(username=request.user).get_full_name()
    Usus = User.objects.all()
    p = Proyectos.objects.get(nombre=proy).id
    n_expediente = Proyectos.objects.get(nombre=proy).n_expediente
    Entries = Entrada_historial.objects.filter(proyecto=p)
    data = {
        'usuario': usuario,
        'Usus': Usus,
        'proyecto': proy,
        'n_expediente': n_expediente,
        'Entries': Entries,
        'autorizado': autorizado,
    }
    return render(request, 'tareas/entradas_asoc_proyecto.html', data)


@login_required
def crear_entrada(request, proy):
    if request.method == 'POST':
        try:
            Entrada_historial.objects.create(
                fecha=request.POST['fecha'],
                resumen=request.POST['resumen'],
                proyecto=Proyectos.objects.get(
                    nombre=request.POST['proyecto']),
                usuario=request.user)
            if request.FILES:
                ent = Entrada_historial.objects.last()
                ent.adjunto = request.FILES['adjunto']
                ent.save()
            messages.success(request, 'Entrada creada con exito.')
        except Exception as e:
            messages.error(
                request, 'Hubo un error al crear la entrada: ' + str(e) + '.')
        return redirect('entradasAsocProyecto', proy=request.POST['proyecto'])
    else:
        usuario = User.objects.get(username=request.user).get_full_name()
        Usus = User.objects.all()
        Proys = Proyectos.objects.all()
        data = {
            'usuario': usuario,
            'Usus': Usus,
            'proyecto': proy,
            'Proys': Proys,
        }
        return render(request, 'tareas/crear_entrada.html', data)


@login_required
def editar_entrada(request):
    # print(f"El id de la entrada es: {request.POST['id']}")
    try:
        entrada = Entrada_historial.objects.get(id=request.POST['id'])
        # print(f'El id del proyecto es: {entrada.proyecto}')
        proyecto = entrada.proyecto
        if request.POST['fecha'] != entrada.fecha.date():
            entrada.fecha = request.POST['fecha']
        if request.POST['resumen'] != entrada.resumen and request.POST['resumen'] != '':
            entrada.resumen = request.POST['resumen']
        entrada.save()
        messages.success(request, "Entrada editada con exito.")
    except Exception as e:
        messages.error(
            request, "Hubo un error al editar la entrada: " + str(e) + ".")
    return redirect('entradasAsocProyecto', proyecto)


@login_required
def borrar_entrada(request, id):
    try:
        entrada = Entrada_historial.objects.get(id=id)
        proyecto = entrada.proyecto
        entrada.delete()
        messages.success(request, "Entrada borrada con exito.")
    except Exception as e:
        messages.error(
            request, "Hubo un error al borrar la entrada: " + str(e) + ".")
    return redirect('entradasAsocProyecto', proyecto)


def cargar_entradas(request):
    """
    Carga entradas históricas de un proyecto desde un archivo Excel.

    Parameters:
    request (HttpRequest): La solicitud HTTP que contiene el archivo Excel.

    Returns:
    redirect: Redirecciona a la página de historico del proyecto.
    """
    total = 0
    contador = 0
    try:
        if request.method == 'POST' and request.FILES.get('excel_file'):
            # obtengo el nombre del proyecto desde la url
            url_origen = request.META.get('HTTP_REFERER')
            path = urlparse(url_origen).path
            ultimo_elemento = path.split("/")[-1]
            proyecto = ultimo_elemento.replace("%20", " ")
            print(proyecto)
            proy_id = Proyectos.objects.get(nombre=proyecto).id
            excel_file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(values_only=True):
                total += 1
                if total == 1:
                    continue

                # Verificar si la fila está vacía
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    break  # Detener el procesamiento si la fila está vacía

                values = []
                for col in range(1, 3):
                    cell = sheet.cell(row=total, column=col)
                    value = cell.value if cell.value is not None else " "
                    values.append(str(value))
                fecha, informacion = values
                try:
                    Entrada_historial.objects.create(
                        fecha=fecha,
                        usuario=User.objects.get(username=request.user),
                        proyecto=Proyectos.objects.get(id=proy_id),
                        resumen=informacion,
                    )
                    contador += 1
                except Exception as e:
                    messages.error(
                        request, f"Sucedió un error inesperado. Error({e})")
            messages.success(
                request, f'Se han cargado {contador} de {total} entradas correctamente.')
    except Exception as e:
        messages.error(
            request, f'No se han podido cargar las entradas. Error({e})')
    return redirect('/tareas/historico/' + proyecto)
